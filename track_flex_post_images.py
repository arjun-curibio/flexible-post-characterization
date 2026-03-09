"""
Track post deflection across a set of still images.

Uses template matching from track_template.py to calculate the displacement
of a user-selected region in each image relative to the first image.
Results are written to an xlsx spreadsheet.

Usage:
    python track_flex_post_images.py
"""

import os
import sys
import math
import argparse

import cv2 as cv
import numpy as np
import openpyxl

from pathlib import Path

import __main__

if __name__ == "__main__":
    sys.path.insert(0, Path(__file__).parent.as_posix())  # add current directory
    sys.path.insert(0, (Path(__file__).parent / "src").as_posix())  # add src directory
    sys.path.insert(0, (Path(__file__).parent / "libs" / "MantaVision" / "src").as_posix())  # add MantaVision src

if not hasattr(__main__, "__file__"):  # operating out of REPL
    sys.path.insert(0, Path(os.curdir).absolute().as_posix())  # add current directory
    sys.path.insert(0, (Path(os.curdir).absolute() / "src").as_posix())  # add src directory
    sys.path.insert(0, (Path(os.curdir).absolute() / "libs" / "MantaVision" / "src").as_posix())  # add MantaVision src

from libs.MantaVision.src.image_utils import openImage
from libs.MantaVision.src.io_utils import contentsOfDir
from libs.MantaVision.src.track_template import matchResults, userDrawnROI, intensityAdjusted, rotatedImage

IMAGE_EXTENSIONS = [".jpg", ".jpeg", ".png", ".tif", ".tiff", ".bmp"]

IMAGE_DIR = None  # set to None to trigger folder selection dialog
OUTPUT_DIR = None  # if None, will save to <image_dir>/tracked/

MAX_ROTATION_DEFAULT = 10.0  # degrees, for rotation search in track_template.py
ROTATION_INCREMENT = 0.25  # degrees.  A smaller number will be more accurate but slower

MAX_DIAMETER = 250.0  # maximum expected diameter of the circular feature to detect, in pixels (used to set HoughCircles maxRadius)
MIN_DIAMETER = 50.0  # minimum expected diameter of the circular feature to detect, in pixels (used to set HoughCircles minRadius)


def load_images(image_dir: str):
    """
    Load all images from a directory, sorted alphabetically by filename.
    Returns a list of (filename, rgb_image) tuples.
    """
    base_dir, files = contentsOfDir(image_dir, IMAGE_EXTENSIONS)
    if files is None:
        print(f"ERROR: No image files found in {image_dir}")
        sys.exit(1)

    # reconstruct full paths and sort by filename
    image_files = sorted(files, key=lambda f: f[0])
    images = []
    for file_name, file_ext in image_files:
        full_name = file_name + file_ext
        file_path = os.path.join(base_dir, full_name)
        image = openImage(file_path, rgb_required=True)
        if image is None:
            print(f"WARNING: Could not load {full_name}, skipping.")
            continue
        images.append((full_name, image))
        print(f"  Loaded: {full_name}")

    if len(images) < 2:
        print("ERROR: Need at least 2 images to compute deflection.")
        sys.exit(1)

    return images


def select_template(first_image_rgb: np.ndarray):
    """
    Show the first image to the user and let them draw an ROI around the post.
    Returns the grayscale template and the ROI dict.
    """
    print("\nA window will open. Draw a rectangle around the post to track, then press ENTER.")
    roi = userDrawnROI(first_image_rgb, title_text="SELECT POST REGION TO TRACK (then press ENTER)")
    if roi is None:
        print("ERROR: No ROI selected. Exiting.")
        sys.exit(1)

    # extract the grayscale template from the selected region
    gray_image = cv.cvtColor(first_image_rgb, cv.COLOR_BGR2GRAY).astype(np.float32)
    gray_adjusted = intensityAdjusted(gray_image)
    template_gray = gray_adjusted[roi["y_start"] : roi["y_end"], roi["x_start"] : roi["x_end"]]

    print(f"  Template size: {template_gray.shape[1]} x {template_gray.shape[0]} pixels")
    return template_gray, roi


def find_most_circular_object(image_rgb: np.ndarray):
    """
    Find the larger of two nearly concentric circles in the image using HoughCircles.
    Returns a dict with the circle center, radius, and diameter,
    or None if no circle is found.

    Expects two concentric circles (e.g. inner and outer edge of a post cross-section).
    Among detected circles that share approximately the same center, returns the larger one.
    """
    gray = cv.cvtColor(image_rgb, cv.COLOR_BGR2GRAY)
    blurred = cv.GaussianBlur(gray, (9, 9), 2)

    img_height = gray.shape[0]

    # Use a small minDist so concentric (nearly overlapping center) circles are both detected
    circles = cv.HoughCircles(
        blurred,
        cv.HOUGH_GRADIENT,
        dp=1,
        minDist=5,
        param1=100,
        param2=30,
        minRadius=int(MIN_DIAMETER // 2),
        maxRadius=int(MAX_DIAMETER // 2),
    )

    if circles is None:
        return None

    detected = circles[0]  # shape: (N, 3) — each row is (x, y, radius)

    if len(detected) < 2:
        # only one circle found — return it
        best = detected[0]
        center_x, center_y, radius = float(best[0]), float(best[1]), float(best[2])
        return {
            "center_x": center_x,
            "center_y": center_y,
            "radius": radius,
            "diameter": radius * 2.0,
        }

    # Find pairs of concentric circles: centers within one min_radius of each other
    # Then pick the pair and return the larger circle
    best_radius = 0.0
    best_circle = None

    for i in range(len(detected)):
        cx_i, cy_i, r_i = float(detected[i][0]), float(detected[i][1]), float(detected[i][2])
        for j in range(i + 1, len(detected)):
            cx_j, cy_j, r_j = float(detected[j][0]), float(detected[j][1]), float(detected[j][2])

            # check if centers are close (concentric)
            center_dist = math.sqrt((cx_i - cx_j) ** 2 + (cy_i - cy_j) ** 2)
            smaller_radius = min(r_i, r_j)
            if center_dist < smaller_radius:
                # these are concentric — pick the larger one
                larger_r = max(r_i, r_j)
                if larger_r > best_radius:
                    best_radius = larger_r
                    if r_i >= r_j:
                        best_circle = (cx_i, cy_i, r_i)
                    else:
                        best_circle = (cx_j, cy_j, r_j)

    if best_circle is not None:
        center_x, center_y, radius = best_circle
    else:
        # no concentric pair found — fall back to the largest detected circle
        largest_idx = int(np.argmax(detected[:, 2]))
        center_x = float(detected[largest_idx][0])
        center_y = float(detected[largest_idx][1])
        radius = float(detected[largest_idx][2])

    return {
        "center_x": center_x,
        "center_y": center_y,
        "radius": radius,
        "diameter": radius * 2.0,
    }


def build_rotated_search_set(gray_adjusted, max_rotation, rotation_increment, previous_rotation, pivot_x, pivot_y):
    """
    Build a search set of rotated versions of the image for matchResults().
    Mirrors the rotation logic in trackTemplate(): generates rotated copies
    from (previous_rotation - max_rotation) to (previous_rotation + max_rotation)
    in rotation_increment steps.

    If max_rotation is None, returns a single-entry search set with no rotation.
    """
    if max_rotation is None:
        return [{"frame": gray_adjusted, "angle": 0.0}]

    search_set = []
    angle = previous_rotation - max_rotation
    max_angle = previous_rotation + max_rotation + rotation_increment
    while angle < max_angle:
        rotated = rotatedImage(gray_adjusted, angle, pivot_x, pivot_y)
        search_set.append({"frame": rotated, "angle": angle})
        angle += rotation_increment

    return search_set


def track_across_images(images, template_gray, max_rotation=None):
    """
    Run template matching on each image and return a list of result dicts.
    Each dict has: filename, image_rgb, match_x, match_y, match_quality,
    match_rotation, and displacements.

    Args:
        images:         list of (filename, rgb_image) tuples
        template_gray:  intensity-adjusted grayscale template
        max_rotation:   maximum rotation to search in degrees (None = no rotation)
    """
    results = []
    reference_x = None
    reference_y = None
    template_h, template_w = template_gray.shape[:2]
    template_half_w = template_w / 2.0
    template_half_h = template_h / 2.0

    # rotation state carried between images
    best_match_rotation = 0.0
    best_match_origin_x = None
    best_match_origin_y = None

    for i, (filename, image_rgb) in enumerate(images):
        # convert to grayscale and intensity-adjust to match the template
        gray_image = cv.cvtColor(image_rgb, cv.COLOR_BGR2GRAY).astype(np.float32)
        gray_adjusted = intensityAdjusted(gray_image)

        # compute rotation pivot from previous match (center of template region)
        if best_match_origin_x is not None:
            pivot_x = best_match_origin_x + template_half_w
            pivot_y = best_match_origin_y + template_half_h
        else:
            # first image: use image center as pivot
            pivot_x = gray_adjusted.shape[1] / 2.0
            pivot_y = gray_adjusted.shape[0] / 2.0

        search_set = build_rotated_search_set(gray_adjusted, max_rotation, ROTATION_INCREMENT, best_match_rotation, pivot_x, pivot_y)

        match_quality, match_coords, match_rotation = matchResults(search_set, template_gray)

        match_x = match_coords[0]
        match_y = match_coords[1]

        # update rotation state for next image
        best_match_rotation = match_rotation
        best_match_origin_x = match_x
        best_match_origin_y = match_y

        # first image is the reference
        if i == 0:
            reference_x = match_x
            reference_y = match_y

        x_displacement = match_x - reference_x
        y_displacement = match_y - reference_y
        xy_displacement = math.sqrt(x_displacement**2 + y_displacement**2)

        # detect the most circular object in the image
        circle_info = find_most_circular_object(image_rgb)

        results.append(
            {
                "filename": filename,
                "image_rgb": image_rgb,
                "match_x": match_x,
                "match_y": match_y,
                "template_w": template_w,
                "template_h": template_h,
                "match_quality": match_quality,
                "match_rotation": match_rotation,
                "x_displacement": x_displacement,
                "y_displacement": y_displacement,
                "xy_displacement": xy_displacement,
                "circle": circle_info,
            }
        )

        rotation_str = f", rot={match_rotation:.1f}°" if max_rotation is not None else ""
        circle_str = f", diameter={circle_info['diameter']:.1f} px" if circle_info else ", no circle found"
        print(f"  [{i+1}/{len(images)}] {filename}: " f"dx={x_displacement:.2f}, dy={y_displacement:.2f}, " f"dxy={xy_displacement:.2f} px  (match={match_quality:.4f}{rotation_str}{circle_str})")

    return results


def draw_match_region(annotated, result):
    """
    Draw the matched template region on the annotated image.
    Uses a rotated rectangle when rotation is non-zero, otherwise an axis-aligned rectangle.
    """
    match_x = result["match_x"]
    match_y = result["match_y"]
    w = result["template_w"]
    h = result["template_h"]
    rotation = result["match_rotation"]
    rect_color = (0, 255, 0)  # green
    rect_thickness = 3

    if abs(rotation) < 0.01:
        # no meaningful rotation — draw a simple rectangle
        top_left = (int(round(match_x)), int(round(match_y)))
        bottom_right = (int(round(match_x + w)), int(round(match_y + h)))
        cv.rectangle(annotated, top_left, bottom_right, rect_color, rect_thickness)
    else:
        # draw a rotated rectangle to show the matched orientation
        # the four corners of the un-rotated template region
        corners = np.array(
            [
                [match_x, match_y],
                [match_x + w, match_y],
                [match_x + w, match_y + h],
                [match_x, match_y + h],
            ],
            dtype=np.float64,
        )

        # rotate corners around the center of the template region
        center_x = match_x + w / 2.0
        center_y = match_y + h / 2.0
        angle_rad = math.radians(rotation)  # OpenCV rotation convention
        cos_a = math.cos(angle_rad)
        sin_a = math.sin(angle_rad)
        rotated_corners = []
        for cx, cy in corners:
            dx = cx - center_x
            dy = cy - center_y
            rx = center_x + dx * cos_a - dy * sin_a
            ry = center_y + dx * sin_a + dy * cos_a
            rotated_corners.append([int(round(rx)), int(round(ry))])
        rotated_corners = np.array(rotated_corners, dtype=np.int32)
        cv.polylines(annotated, [rotated_corners], isClosed=True, color=rect_color, thickness=rect_thickness)


def save_annotated_images(results, output_dir: str):
    """
    For each result, draw a rectangle where the template was found and
    overlay the displacement text, then save the annotated image.
    """
    os.makedirs(output_dir, exist_ok=True)

    for i, result in enumerate(results):
        annotated = result["image_rgb"].copy()

        # pick a font size relative to image width so text is readable
        img_width = annotated.shape[1]
        font_scale = max(0.5, img_width / 1500.0)
        font_thickness = max(1, int(font_scale * 2))
        font = cv.FONT_HERSHEY_SIMPLEX

        # draw the matched template region (axis-aligned or rotated)
        draw_match_region(annotated, result)

        # draw the detected circle
        circle = result.get("circle")
        if circle is not None:
            center = (int(round(circle["center_x"])), int(round(circle["center_y"])))
            radius = int(round(circle["radius"]))
            cv.circle(annotated, center, radius, (0, 255, 255), 2)  # yellow circle
            cv.circle(annotated, center, 3, (0, 255, 255), -1)  # center dot

            # diameter label near the circle
            circle_label = f"d={circle['diameter']:.1f} px"
            cl_size, _ = cv.getTextSize(circle_label, font, font_scale, font_thickness)
            cl_x = center[0] - cl_size[0] // 2
            cl_y = center[1] + radius + cl_size[1] + 10
            # keep label within image bounds
            cl_x = max(0, min(cl_x, img_width - cl_size[0]))
            cl_y = min(cl_y, annotated.shape[0] - 4)
            bg_tl = (cl_x - 2, cl_y - cl_size[1] - 4)
            bg_br = (cl_x + cl_size[0] + 2, cl_y + 4)
            cv.rectangle(annotated, bg_tl, bg_br, (0, 0, 0), cv.FILLED)
            cv.putText(annotated, circle_label, (cl_x, cl_y), font, font_scale, (0, 255, 255), font_thickness)

        # build the displacement label
        dx = result["x_displacement"]
        dy = result["y_displacement"]
        dxy = result["xy_displacement"]
        rotation = result["match_rotation"]
        label = f"dx={dx:.1f}  dy={dy:.1f}  dxy={dxy:.1f} px"
        if abs(rotation) >= 0.01:
            label += f"  rot={rotation:.1f} deg"

        # position the label just above the match region (or below if no room)
        top_left_y = int(round(result["match_y"]))
        bottom_right_y = int(round(result["match_y"] + result["template_h"]))
        text_size, _ = cv.getTextSize(label, font, font_scale, font_thickness)
        text_x = int(round(result["match_x"]))
        text_y = top_left_y - 10
        if text_y - text_size[1] < 0:
            text_y = bottom_right_y + text_size[1] + 10

        # draw a dark background behind the text for contrast
        bg_top_left = (text_x - 2, text_y - text_size[1] - 4)
        bg_bottom_right = (text_x + text_size[0] + 2, text_y + 4)
        cv.rectangle(annotated, bg_top_left, bg_bottom_right, (0, 0, 0), cv.FILLED)
        cv.putText(annotated, label, (text_x, text_y), font, font_scale, (0, 255, 0), font_thickness)

        # save the annotated image
        base_name, ext = os.path.splitext(result["filename"])
        output_filename = f"{base_name}_tracked{ext}"
        output_path = os.path.join(output_dir, output_filename)
        cv.imwrite(output_path, annotated)
        print(f"  Saved: {output_filename}")

    print(f"\nAnnotated images saved to: {output_dir}")


def write_results_to_xlsx(results, output_path: str):
    """
    Write the deflection results to an xlsx spreadsheet.
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    if sheet is None:
        print("ERROR: Could not create Excel workbook.")
        return

    sheet.title = "Deflection Results"

    # column headings
    headings = [
        "Image Name",
        "X Displacement (px)",
        "Y Displacement (px)",
        "XY Displacement (px)",
        "Rotation (deg)",
        "Circle Diameter (px)",
        "Match Quality",
        "Match X (px)",
        "Match Y (px)",
    ]
    for col, heading in enumerate(headings, start=1):
        sheet.cell(row=1, column=col, value=heading)

    # data rows
    for row_idx, result in enumerate(results, start=2):
        sheet.cell(row=row_idx, column=1, value=result["filename"])
        sheet.cell(row=row_idx, column=2, value=round(result["x_displacement"], 4))
        sheet.cell(row=row_idx, column=3, value=round(result["y_displacement"], 4))
        sheet.cell(row=row_idx, column=4, value=round(result["xy_displacement"], 4))
        sheet.cell(row=row_idx, column=5, value=round(result["match_rotation"], 4))
        circle = result.get("circle")
        if circle is not None:
            sheet.cell(row=row_idx, column=6, value=round(circle["diameter"], 4))
        else:
            sheet.cell(row=row_idx, column=6, value="N/A")
        sheet.cell(row=row_idx, column=7, value=round(result["match_quality"], 6))
        sheet.cell(row=row_idx, column=8, value=round(result["match_x"], 4))
        sheet.cell(row=row_idx, column=9, value=round(result["match_y"], 4))

    # auto-size columns for readability
    for col in sheet.columns:
        max_length = 0
        col_letter = col[0].column_letter  # type: ignore
        for cell in col:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[col_letter].width = max_length + 2

    workbook.save(output_path)
    print(f"\nResults saved to: {output_path}")


if __name__ == "__main__":
    if IMAGE_DIR is None:
        # open filedialog using PyQt6
        from PyQt6.QtWidgets import QApplication, QFileDialog

        app = QApplication([])
        selected_dir = QFileDialog.getExistingDirectory(None, "Select folder containing images to analyze")
        if not selected_dir:
            print("ERROR: No folder selected. Exiting.")
            sys.exit(1)
        IMAGE_DIR = selected_dir

    if not os.path.isdir(IMAGE_DIR):
        print(f"ERROR: '{IMAGE_DIR}' is not a valid directory.")
        sys.exit(1)

    output_path = OUTPUT_DIR
    if output_path is None:
        output_path = os.path.join(Path(IMAGE_DIR).parent, f"{Path(IMAGE_DIR).name}_deflection_results.xlsx")

    print(f"Loading images from: {IMAGE_DIR}")
    images = load_images(IMAGE_DIR)
    print(f"Found {len(images)} images.\n")

    template_gray, roi = select_template(images[0][1])

    max_rotation = MAX_ROTATION_DEFAULT
    if max_rotation is not None:
        print(f"Rotation search enabled: +/- {max_rotation} degrees in 0.5 degree steps")

    print("\nTracking template across all images...")
    results = track_across_images(images, template_gray, max_rotation=max_rotation)

    print("\nSaving annotated images...")
    annotated_dir = os.path.join(IMAGE_DIR, "tracked")
    save_annotated_images(results, annotated_dir)

    write_results_to_xlsx(results, output_path)

    print("\nDone.")
