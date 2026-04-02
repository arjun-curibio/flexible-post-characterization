"""
Track y-displacement of white circular objects across a set of still images.

Uses Hough circle detection to find white filled circles within user-specified
diameter bounds. The starting (reference) position is always the highest
detected position (smallest y-coordinate). Results are written to an xlsx
spreadsheet and annotated images are saved.

Usage:
    python track_white_circles.py
"""

import os
import sys
import math

import cv2
import numpy as np
import openpyxl

from pathlib import Path

IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".tif", ".tiff", ".bmp"}

# ── User Configuration ────────────────────────────────────────────────────────
IMAGE_DIR = None  # set to None to trigger folder selection dialog
OUTPUT_DIR = None  # if None, will save to <IMAGE_DIR>/tracked/

MIN_DIAMETER = 200  # minimum expected diameter of white circles, in pixels
MAX_DIAMETER = 500  # maximum expected diameter of white circles, in pixels

# White threshold: pixels above this value (0-255) are considered "white"
WHITE_THRESHOLD = 100
# Minimum fraction of the circle area that must be white to count as a white filled circle
WHITE_FILL_RATIO = 0.5

DIAM = 1000  # known physical diameter of the circular object, in microns
WEIGHT = 80  # weight applied to the post, in milligrams
GRAVITY = 9.80665  # m/s^2



# ──────────────────────────────────────────────────────────────────────────────


def load_images(image_dir: str):
    """Load all images from a directory, sorted alphabetically by filename."""
    entries = sorted(os.listdir(image_dir))
    images = []
    for entry in entries:
        ext = os.path.splitext(entry)[1].lower()
        if ext not in IMAGE_EXTENSIONS:
            continue
        file_path = os.path.join(image_dir, entry)
        if not os.path.isfile(file_path):
            continue
        image = cv2.imread(file_path, cv2.IMREAD_COLOR)
        if image is None:
            print(f"WARNING: Could not load {entry}, skipping.")
            continue
        images.append((entry, image))
        print(f"  Loaded: {entry}")

    if len(images) < 2:
        print("ERROR: Need at least 2 images to compute displacement.")
        sys.exit(1)

    return images


def circle_white_fill(gray_image: np.ndarray, cx: float, cy: float, radius: float, threshold: int) -> float:
    """Return the fraction of pixels inside the circle that are above the white threshold."""
    h, w = gray_image.shape[:2]
    ix, iy, ir = int(round(cx)), int(round(cy)), int(round(radius))

    # bounding box clipped to image
    x0 = max(ix - ir, 0)
    x1 = min(ix + ir + 1, w)
    y0 = max(iy - ir, 0)
    y1 = min(iy + ir + 1, h)

    if x1 <= x0 or y1 <= y0:
        return 0.0

    # create a circular mask
    yy, xx = np.ogrid[y0:y1, x0:x1]
    mask = ((xx - cx) ** 2 + (yy - cy) ** 2) <= (radius**2)

    total_pixels = np.count_nonzero(mask)
    if total_pixels == 0:
        return 0.0

    roi = gray_image[y0:y1, x0:x1]
    white_pixels = np.count_nonzero((roi > threshold) & mask)
    return white_pixels / total_pixels


def find_white_circles(image_rgb: np.ndarray):
    """
    Find all white filled circles in the image within the configured diameter bounds.

    Returns a list of dicts, each with center_x, center_y, radius, diameter, and white_fill.
    The list is sorted by white_fill descending (best match first).
    """
    gray = cv2.cvtColor(image_rgb, cv2.COLOR_BGR2GRAY)
    blurred = cv2.GaussianBlur(gray, (9, 9), 2)

    min_radius = int(MIN_DIAMETER // 2)
    max_radius = int(MAX_DIAMETER // 2)

    circles = cv2.HoughCircles(
        blurred,
        cv2.HOUGH_GRADIENT,
        dp=1,
        minDist=min_radius,
        param1=100,
        param2=30,
        minRadius=min_radius,
        maxRadius=max_radius,
    )

    if circles is None:
        return []

    detected = circles[0]
    results = []
    for det in detected:
        cx, cy, r = float(det[0]), float(det[1]), float(det[2])
        diameter = r * 2.0
        if diameter < MIN_DIAMETER or diameter > MAX_DIAMETER:
            continue

        fill = circle_white_fill(gray, cx, cy, r, WHITE_THRESHOLD)
        if fill >= WHITE_FILL_RATIO:
            results.append(
                {
                    "center_x": cx,
                    "center_y": cy,
                    "radius": r,
                    "diameter": diameter,
                    "white_fill": fill,
                }
            )

    results.sort(key=lambda c: c["white_fill"], reverse=True)
    return results


def track_across_images(images):
    """
    Detect white circles in each image and compute y-displacement.
    The reference y is the highest position (smallest y) across all images.

    Returns a list of result dicts, one per image.
    """
    # first pass: detect circles in every image
    all_detections = []
    for i, (filename, image_rgb) in enumerate(images):
        circles = find_white_circles(image_rgb)
        all_detections.append((filename, image_rgb, circles))
        count = len(circles)
        if count > 0:
            best = circles[0]
            print(
                f"  [{i+1}/{len(images)}] {filename}: {count} white circle(s) found, "
                f"best at ({best['center_x']:.1f}, {best['center_y']:.1f}), "
                f"d={best['diameter']:.1f} px, fill={best['white_fill']:.1%}"
            )
        else:
            print(f"  [{i+1}/{len(images)}] {filename}: no white circles found")

    # determine reference y: the highest position (smallest center_y) of the best circle
    reference_y = None
    for _, _, circles in all_detections:
        if circles:
            cy = circles[0]["center_y"]
            if reference_y is None or cy < reference_y:
                reference_y = cy

    if reference_y is None:
        print("ERROR: No white circles detected in any image.")
        # sys.exit(1)
        return []

    # second pass: build results with displacement, microns/pixel, and force
    # force = weight (mg -> kg) * gravity => Newtons, convert to mN
    force_mN = (WEIGHT / 1e6) * GRAVITY * 1e3  # millinewtons

    results = []
    for filename, image_rgb, circles in all_detections:
        best_circle = circles[0] if circles else None
        if best_circle is not None:
            y_displacement_px = best_circle["center_y"] - reference_y
            microns_per_pixel = DIAM / best_circle["diameter"]
            y_displacement_um = y_displacement_px * microns_per_pixel
            # stiffness k = F / d, force in N, displacement in m
            if y_displacement_um > 0:
                force_N = (WEIGHT / 1e6) * GRAVITY
                displacement_m = y_displacement_um / 1e6
                stiffness_N_m = force_N / displacement_m
            else:
                stiffness_N_m = None
        else:
            y_displacement_px = None
            microns_per_pixel = None
            y_displacement_um = None
            stiffness_N_m = None

        results.append(
            {
                "filename": filename,
                "image_rgb": image_rgb,
                "circles": circles,
                "best_circle": best_circle,
                "y_displacement_px": y_displacement_px,
                "y_displacement_um": y_displacement_um,
                "microns_per_pixel": microns_per_pixel,
                "force_mN": force_mN,
                "stiffness_N_m": stiffness_N_m,
                "reference_y": reference_y,
            }
        )

    return results


def save_annotated_images(results, output_dir: str):
    """Draw detected circles and displacement text on each image, then save."""
    os.makedirs(output_dir, exist_ok=True)

    for result in results:
        annotated = result["image_rgb"].copy()
        img_width = annotated.shape[1]
        font_scale = max(0.5, img_width / 1500.0)
        font_thickness = max(1, int(font_scale * 2))
        font = cv2.FONT_HERSHEY_SIMPLEX

        # draw all detected white circles
        for j, circle in enumerate(result["circles"]):
            center = (int(round(circle["center_x"])), int(round(circle["center_y"])))
            radius = int(round(circle["radius"]))

            if j == 0:
                # best circle: yellow
                color = (0, 255, 255)
                thickness = 3
            else:
                # other circles: cyan
                color = (255, 255, 0)
                thickness = 2

            cv2.circle(annotated, center, radius, color, thickness)
            cv2.circle(annotated, center, 3, color, -1)

            # diameter label
            label = f"d={circle['diameter']:.1f} px"
            ls, _ = cv2.getTextSize(label, font, font_scale * 0.8, font_thickness)
            lx = max(0, min(center[0] - ls[0] // 2, img_width - ls[0]))
            ly = min(center[1] + radius + ls[1] + 10, annotated.shape[0] - 4)
            cv2.rectangle(annotated, (lx - 2, ly - ls[1] - 4), (lx + ls[0] + 2, ly + 4), (0, 0, 0), cv2.FILLED)
            cv2.putText(annotated, label, (lx, ly), font, font_scale * 0.8, color, font_thickness)

        # displacement label for the best circle
        best = result["best_circle"]
        y_disp_px = result["y_displacement_px"]
        y_disp_um = result["y_displacement_um"]
        if best is not None and y_disp_px is not None:
            disp_label = f"dy={y_disp_px:.1f} px ({y_disp_um:.1f} um)"
            center = (int(round(best["center_x"])), int(round(best["center_y"])))
            ts, _ = cv2.getTextSize(disp_label, font, font_scale, font_thickness)
            tx = max(0, min(center[0] - ts[0] // 2, img_width - ts[0]))
            ty = max(ts[1] + 10, center[1] - int(round(best["radius"])) - 10)
            cv2.rectangle(annotated, (tx - 2, ty - ts[1] - 4), (tx + ts[0] + 2, ty + 4), (0, 0, 0), cv2.FILLED)
            cv2.putText(annotated, disp_label, (tx, ty), font, font_scale, (0, 255, 0), font_thickness)

        # draw reference line
        ref_y = int(round(result["reference_y"]))
        cv2.line(annotated, (0, ref_y), (img_width, ref_y), (0, 0, 255), 1, cv2.LINE_AA)

        base_name, ext = os.path.splitext(result["filename"])
        output_filename = f"{base_name}_tracked{ext}"
        output_path = os.path.join(output_dir, output_filename)
        cv2.imwrite(output_path, annotated)
        print(f"  Saved: {output_filename}")

    print(f"\nAnnotated images saved to: {output_dir}")


def write_results_to_xlsx(results, output_path: str):
    """Write results to an xlsx spreadsheet."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    if sheet is None:
        print("ERROR: Could not create Excel workbook.")
        return

    sheet.title = "White Circle Tracking"

    headings = [
        "Image Name",
        "Y Displacement (px)",
        "Y Displacement (um)",
        "Microns/Pixel",
        "Force (mN)",
        "Stiffness (N/m)",
        "Circle Center X (px)",
        "Circle Center Y (px)",
        "Circle Diameter (px)",
        "White Fill (%)",
        "Circles Found",
        "Reference Y (px)",
    ]
    for col, heading in enumerate(headings, start=1):
        sheet.cell(row=1, column=col, value=heading)

    for row_idx, result in enumerate(results, start=2):
        sheet.cell(row=row_idx, column=1, value=result["filename"])

        best = result["best_circle"]
        y_disp_px = result["y_displacement_px"]
        y_disp_um = result["y_displacement_um"]
        um_per_px = result["microns_per_pixel"]
        stiffness = result["stiffness_N_m"]

        if y_disp_px is not None:
            sheet.cell(row=row_idx, column=2, value=round(y_disp_px, 4))
            sheet.cell(row=row_idx, column=3, value=round(y_disp_um, 4))
            sheet.cell(row=row_idx, column=4, value=round(um_per_px, 4))
            sheet.cell(row=row_idx, column=5, value=round(result["force_mN"], 6))
            sheet.cell(row=row_idx, column=6, value=round(stiffness, 4) if stiffness is not None else "N/A")
        else:
            for c in range(2, 7):
                sheet.cell(row=row_idx, column=c, value="N/A")

        if best is not None:
            sheet.cell(row=row_idx, column=7, value=round(best["center_x"], 4))
            sheet.cell(row=row_idx, column=8, value=round(best["center_y"], 4))
            sheet.cell(row=row_idx, column=9, value=round(best["diameter"], 4))
            sheet.cell(row=row_idx, column=10, value=round(best["white_fill"] * 100, 2))
        else:
            for c in range(7, 11):
                sheet.cell(row=row_idx, column=c, value="N/A")

        sheet.cell(row=row_idx, column=11, value=len(result["circles"]))
        sheet.cell(row=row_idx, column=12, value=round(result["reference_y"], 4))

    for col in sheet.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[col_letter].width = max_length + 2

    workbook.save(output_path)
    print(f"\nResults saved to: {output_path}")


if __name__ == "__main__":
    if IMAGE_DIR is None:
        from PyQt6.QtWidgets import QApplication, QFileDialog

        app = QApplication([])
        selected_dir = QFileDialog.getExistingDirectory(None, "Select folder containing images to analyze")
        if not selected_dir:
            print("ERROR: No folder selected. Exiting.")
            sys.exit(1)
        IMAGE_DIR = selected_dir

    if not os.path.isdir(IMAGE_DIR):
        print(f"ERROR: '{IMAGE_DIR}' is not a valid directory.")
        sys.exit(1)

    output_dir = OUTPUT_DIR
    if output_dir is None:
        output_dir = os.path.join(IMAGE_DIR, "tracked")

    output_xlsx = os.path.join(Path(IMAGE_DIR).parent, f"{Path(IMAGE_DIR).name}_white_circle_results.xlsx")

    print(f"Loading images from: {IMAGE_DIR}")
    print(f"Diameter bounds: {MIN_DIAMETER:.0f} - {MAX_DIAMETER:.0f} px")
    print(f"White threshold: {WHITE_THRESHOLD}, min fill ratio: {WHITE_FILL_RATIO:.0%}\n")

    images = load_images(IMAGE_DIR)
    print(f"Found {len(images)} images.\n")

    print("Detecting white circles and computing y-displacement...")
    results = track_across_images(images)

    print("\nSaving annotated images...")
    save_annotated_images(results, output_dir)

    write_results_to_xlsx(results, output_xlsx)

    print("\nDone.")
