"""
Characterize flexible post spring constant from a set of still images.

Part 1: Track post deflection across images using template matching.
         Writes displacement results to an xlsx spreadsheet and saves
         annotated images.
Part 2: Compute spring constant from the deflection data using linear
         regression of force vs. displacement.

Usage:
    python characterize_flex_post.py
"""

import os
import sys
import math

import cv2
import numpy as np
import openpyxl
import matplotlib.pyplot as plt
from scipy.stats import linregress, shapiro
from pathlib import Path

import __main__

if __name__ == "__main__":
    sys.path.insert(0, Path(__file__).parent.as_posix())
    sys.path.insert(0, (Path(__file__).parent / "src").as_posix())
    sys.path.insert(0, (Path(__file__).parent / "libs" / "MantaVision" / "src").as_posix())

if not hasattr(__main__, "__file__"):
    sys.path.insert(0, Path(os.curdir).absolute().as_posix())
    sys.path.insert(0, (Path(os.curdir).absolute() / "src").as_posix())
    sys.path.insert(0, (Path(os.curdir).absolute() / "libs" / "MantaVision" / "src").as_posix())

from libs.MantaVision.src.image_utils import openImage
from libs.MantaVision.src.track_template import matchResults, userDrawnROI, intensityAdjusted, rotatedImage


def contentsOfDir(dir_path, search_terms, search_extension_only=True):
    """Return (base_dir, [(file_name, file_ext), ...]) for files matching search_terms."""
    import glob as _glob

    all_files_found = []
    if os.path.isdir(dir_path):
        base_dir = dir_path
        for term in search_terms:
            pattern = "*" + term if search_extension_only else "*" + term + "*"
            all_files_found.extend(_glob.glob(os.path.join(dir_path, pattern)))
    else:
        base_dir = os.path.dirname(dir_path)
        all_files_found = [dir_path]
    if not all_files_found:
        return None, None
    files = []
    for fp in all_files_found:
        name, ext = os.path.splitext(os.path.basename(fp))
        files.append((name, ext))
    return base_dir, files


# ============================================================================
# USER CONFIGURATION
# ============================================================================

SCALE = None  # pixels per mm — set to None to calibrate interactively by drawing a circle
FEATURE_DIAMETER_MM = 1.0  # known diameter of the reference feature in mm (used during calibration)

WEIGHT_MASS = 0.000050  # mass of each individual weight (kg)
WEIGHT_NUM = np.arange(6)  # weight indices (0, 1, 2, 3, 4, 5) — one image per weight level
GRAVITY = 9.81  # m/s^2

IMAGE_DIR = None  # set to None to trigger folder selection dialog
OUTPUT_DIR = None  # if None, will save xlsx next to IMAGE_DIR

MAX_ROTATION_DEFAULT = 10.0  # degrees, for rotation search
ROTATION_INCREMENT = 0.25  # degrees — smaller is more accurate but slower

# ============================================================================
# SCALE CALIBRATION — INTERACTIVE CIRCLE DRAWING
# ============================================================================

IMAGE_EXTENSIONS = [".jpg", ".jpeg", ".png", ".tif", ".tiff", ".bmp"]


def user_drawn_circle(image_rgb: np.ndarray, title_text: str = "DRAW CIRCLE around reference feature, then press ENTER"):
    """
    Let the user draw a circle on the image by clicking one edge of the feature,
    dragging across the center to the opposite edge, and releasing.
    The circle is defined by the two edge points (endpoints of a diameter).
    Press ENTER to confirm, ESC to cancel.

    Returns a dict with center_x, center_y, radius, diameter (all in pixels),
    or None if cancelled.
    """
    state = {"start": None, "end": None, "dragging": False}

    def mouse_callback(event, x, y, flags, param):
        if event == cv2.EVENT_LBUTTONDOWN:
            state["start"] = (x, y)
            state["end"] = (x, y)
            state["dragging"] = True
        elif event == cv2.EVENT_MOUSEMOVE and state["dragging"]:
            state["end"] = (x, y)
        elif event == cv2.EVENT_LBUTTONUP:
            state["end"] = (x, y)
            state["dragging"] = False

    window_name = title_text
    window_flags = cv2.WINDOW_KEEPRATIO | cv2.WINDOW_NORMAL | cv2.WINDOW_GUI_NORMAL
    cv2.namedWindow(window_name, flags=window_flags)
    cv2.resizeWindow(window_name, width=1280, height=720)
    cv2.moveWindow(window_name, x=200, y=200)
    cv2.setMouseCallback(window_name, mouse_callback)

    while True:
        frame = image_rgb.copy()
        if state["start"] is not None and state["end"] is not None:
            sx, sy = state["start"]
            ex, ey = state["end"]
            cx = (sx + ex) / 2.0
            cy = (sy + ey) / 2.0
            radius = math.sqrt((ex - sx) ** 2 + (ey - sy) ** 2) / 2.0

            if radius > 0:
                cv2.circle(frame, (int(cx), int(cy)), int(radius), (0, 255, 0), 2)
                cv2.circle(frame, (int(cx), int(cy)), 3, (0, 255, 0), -1)
                # show diameter text
                diameter_px = radius * 2.0
                label = f"diameter = {diameter_px:.1f} px"
                font = cv2.FONT_HERSHEY_SIMPLEX
                img_width = frame.shape[1]
                font_scale = max(0.5, img_width / 1500.0)
                font_thickness = max(1, int(font_scale * 2))
                text_pos = (int(cx + radius) + 10, int(cy))
                cv2.putText(frame, label, text_pos, font, font_scale, (0, 255, 0), font_thickness)

        cv2.imshow(window_name, frame)
        key = cv2.waitKey(30) & 0xFF

        if key == 13:  # ENTER
            break
        elif key == 27:  # ESC
            cv2.destroyAllWindows()
            return None

    cv2.destroyAllWindows()

    if state["start"] is None or state["end"] is None:
        return None

    sx, sy = state["start"]
    ex, ey = state["end"]
    cx = (sx + ex) / 2.0
    cy = (sy + ey) / 2.0
    radius = math.sqrt((ex - sx) ** 2 + (ey - sy) ** 2) / 2.0

    if radius <= 0:
        return None

    return {
        "center_x": cx,
        "center_y": cy,
        "radius": radius,
        "diameter": radius * 2.0,
    }


def calibrate_scale(image_rgb: np.ndarray, feature_diameter_mm: float):
    """
    Let the user draw a circle on the image to match a reference feature
    of known diameter. Returns scale in pixels per meter.
    """
    print(f"\nA window will open. Draw a circle matching the {feature_diameter_mm} mm reference feature, then press ENTER.")
    circle = user_drawn_circle(image_rgb)
    if circle is None:
        print("ERROR: No circle drawn. Exiting.")
        sys.exit(1)

    diameter_px = circle["diameter"]
    scale_px_per_mm = diameter_px / feature_diameter_mm
    scale_px_per_m = scale_px_per_mm * 1000.0

    print(f"  Detected diameter: {diameter_px:.1f} px")
    print(f"  Scale: {scale_px_per_mm:.1f} px/mm  ({scale_px_per_m:.1f} px/m)")

    return scale_px_per_m


# ============================================================================
# PART 1 — IMAGE TRACKING FUNCTIONS
# ============================================================================


def load_images(image_dir: str):
    """
    Load all images from a directory, sorted alphabetically by filename.
    Returns a list of (filename, rgb_image) tuples.
    """
    base_dir, files = contentsOfDir(image_dir, IMAGE_EXTENSIONS)
    if files is None:
        print(f"ERROR: No image files found in {image_dir}")
        sys.exit(1)

    image_files = sorted(files, key=lambda f: f[0])
    images = []
    for file_name, file_ext in image_files:
        full_name = file_name + file_ext
        file_path = os.path.join(base_dir, full_name)
        image = openImage(file_path, rgb_required=True)
        if image is None:
            print(f"WARNING: Could not load {full_name}, skipping.")
            continue
        images.append((full_name, image))
        print(f"  Loaded: {full_name}")

    if len(images) < 2:
        print("ERROR: Need at least 2 images to compute deflection.")
        sys.exit(1)

    return images


def select_template(first_image_rgb: np.ndarray):
    """
    Show the first image to the user and let them draw an ROI around the post.
    Returns the grayscale template and the ROI dict.
    """
    print("\nA window will open. Draw a rectangle around the post to track, then press ENTER.")
    roi = userDrawnROI(first_image_rgb, title_text="SELECT POST REGION TO TRACK (then press ENTER)")
    if roi is None:
        print("ERROR: No ROI selected. Exiting.")
        sys.exit(1)

    gray_image = cv2.cvtColor(first_image_rgb, cv2.COLOR_BGR2GRAY).astype(np.float32)
    gray_adjusted = intensityAdjusted(gray_image)
    template_gray = gray_adjusted[roi["y_start"] : roi["y_end"], roi["x_start"] : roi["x_end"]]

    print(f"  Template size: {template_gray.shape[1]} x {template_gray.shape[0]} pixels")
    return template_gray, roi


def build_rotated_search_set(gray_adjusted, max_rotation, rotation_increment, previous_rotation, pivot_x, pivot_y):
    """
    Build a search set of rotated versions of the image for matchResults().
    """
    if max_rotation is None:
        return [{"frame": gray_adjusted, "angle": 0.0}]

    search_set = []
    angle = previous_rotation - max_rotation
    max_angle = previous_rotation + max_rotation + rotation_increment
    while angle < max_angle:
        rotated = rotatedImage(gray_adjusted, angle, pivot_x, pivot_y)
        search_set.append({"frame": rotated, "angle": angle})
        angle += rotation_increment

    return search_set


def track_across_images(images, template_gray, max_rotation=None):
    """
    Run template matching on each image and return a list of result dicts.
    """
    results = []
    reference_x = None
    reference_y = None
    template_h, template_w = template_gray.shape[:2]
    template_half_w = template_w / 2.0
    template_half_h = template_h / 2.0

    best_match_rotation = 0.0
    best_match_origin_x = None
    best_match_origin_y = None

    for i, (filename, image_rgb) in enumerate(images):
        gray_image = cv2.cvtColor(image_rgb, cv2.COLOR_BGR2GRAY).astype(np.float32)
        gray_adjusted = intensityAdjusted(gray_image)

        if best_match_origin_x is not None:
            pivot_x = best_match_origin_x + template_half_w
            pivot_y = best_match_origin_y + template_half_h
        else:
            pivot_x = gray_adjusted.shape[1] / 2.0
            pivot_y = gray_adjusted.shape[0] / 2.0

        search_set = build_rotated_search_set(gray_adjusted, max_rotation, ROTATION_INCREMENT, best_match_rotation, pivot_x, pivot_y)

        match_quality, match_coords, match_rotation = matchResults(search_set, template_gray)

        match_x = match_coords[0]
        match_y = match_coords[1]

        best_match_rotation = match_rotation
        best_match_origin_x = match_x
        best_match_origin_y = match_y

        if i == 0:
            reference_x = match_x
            reference_y = match_y

        x_displacement = match_x - reference_x
        y_displacement = match_y - reference_y
        xy_displacement = math.sqrt(x_displacement**2 + y_displacement**2)

        results.append(
            {
                "filename": filename,
                "image_rgb": image_rgb,
                "match_x": match_x,
                "match_y": match_y,
                "template_w": template_w,
                "template_h": template_h,
                "match_quality": match_quality,
                "match_rotation": match_rotation,
                "x_displacement": x_displacement,
                "y_displacement": y_displacement,
                "xy_displacement": xy_displacement,
            }
        )

        rotation_str = f", rot={match_rotation:.1f}\u00b0" if max_rotation is not None else ""
        print(f"  [{i+1}/{len(images)}] {filename}: " f"dx={x_displacement:.2f}, dy={y_displacement:.2f}, " f"dxy={xy_displacement:.2f} px  (match={match_quality:.4f}{rotation_str})")

    return results


def draw_match_region(annotated, result):
    """
    Draw the matched template region on the annotated image.
    """
    match_x = result["match_x"]
    match_y = result["match_y"]
    w = result["template_w"]
    h = result["template_h"]
    rotation = result["match_rotation"]
    rect_color = (0, 255, 0)
    rect_thickness = 3

    if abs(rotation) < 0.01:
        top_left = (int(round(match_x)), int(round(match_y)))
        bottom_right = (int(round(match_x + w)), int(round(match_y + h)))
        cv2.rectangle(annotated, top_left, bottom_right, rect_color, rect_thickness)
    else:
        corners = np.array(
            [
                [match_x, match_y],
                [match_x + w, match_y],
                [match_x + w, match_y + h],
                [match_x, match_y + h],
            ],
            dtype=np.float64,
        )

        center_x = match_x + w / 2.0
        center_y = match_y + h / 2.0
        angle_rad = math.radians(rotation)
        cos_a = math.cos(angle_rad)
        sin_a = math.sin(angle_rad)
        rotated_corners = []
        for cx, cy in corners:
            dx = cx - center_x
            dy = cy - center_y
            rx = center_x + dx * cos_a - dy * sin_a
            ry = center_y + dx * sin_a + dy * cos_a
            rotated_corners.append([int(round(rx)), int(round(ry))])
        rotated_corners = np.array(rotated_corners, dtype=np.int32)
        cv2.polylines(annotated, [rotated_corners], isClosed=True, color=rect_color, thickness=rect_thickness)


def save_annotated_images(results, output_dir: str):
    """
    Draw a rectangle where the template was found, overlay displacement text,
    and save the annotated image.
    """
    os.makedirs(output_dir, exist_ok=True)

    for i, result in enumerate(results):
        annotated = result["image_rgb"].copy()

        img_width = annotated.shape[1]
        font_scale = max(0.5, img_width / 1500.0)
        font_thickness = max(1, int(font_scale * 2))
        font = cv2.FONT_HERSHEY_SIMPLEX

        draw_match_region(annotated, result)

        dx = result["x_displacement"]
        dy = result["y_displacement"]
        dxy = result["xy_displacement"]
        rotation = result["match_rotation"]
        label = f"dx={dx:.1f}  dy={dy:.1f}  dxy={dxy:.1f} px"
        if abs(rotation) >= 0.01:
            label += f"  rot={rotation:.1f} deg"

        top_left_y = int(round(result["match_y"]))
        bottom_right_y = int(round(result["match_y"] + result["template_h"]))
        text_size, _ = cv2.getTextSize(label, font, font_scale, font_thickness)
        text_x = int(round(result["match_x"]))
        text_y = top_left_y - 10
        if text_y - text_size[1] < 0:
            text_y = bottom_right_y + text_size[1] + 10

        bg_top_left = (text_x - 2, text_y - text_size[1] - 4)
        bg_bottom_right = (text_x + text_size[0] + 2, text_y + 4)
        cv2.rectangle(annotated, bg_top_left, bg_bottom_right, (0, 0, 0), cv2.FILLED)
        cv2.putText(annotated, label, (text_x, text_y), font, font_scale, (0, 255, 0), font_thickness)

        base_name, ext = os.path.splitext(result["filename"])
        output_filename = f"{base_name}_tracked{ext}"
        output_path = os.path.join(output_dir, output_filename)
        cv2.imwrite(output_path, annotated)
        print(f"  Saved: {output_filename}")

    print(f"\nAnnotated images saved to: {output_dir}")


def write_results_to_xlsx(results, output_path: str):
    """
    Write the deflection results to an xlsx spreadsheet.
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    if sheet is None:
        print("ERROR: Could not create Excel workbook.")
        return

    sheet.title = "Deflection Results"

    headings = [
        "Image Name",
        "X Displacement (px)",
        "Y Displacement (px)",
        "XY Displacement (px)",
        "Rotation (deg)",
        "Match Quality",
        "Match X (px)",
        "Match Y (px)",
    ]
    for col, heading in enumerate(headings, start=1):
        sheet.cell(row=1, column=col, value=heading)

    for row_idx, result in enumerate(results, start=2):
        sheet.cell(row=row_idx, column=1, value=result["filename"])
        sheet.cell(row=row_idx, column=2, value=round(result["x_displacement"], 4))
        sheet.cell(row=row_idx, column=3, value=round(result["y_displacement"], 4))
        sheet.cell(row=row_idx, column=4, value=round(result["xy_displacement"], 4))
        sheet.cell(row=row_idx, column=5, value=round(result["match_rotation"], 4))
        sheet.cell(row=row_idx, column=6, value=round(result["match_quality"], 6))
        sheet.cell(row=row_idx, column=7, value=round(result["match_x"], 4))
        sheet.cell(row=row_idx, column=8, value=round(result["match_y"], 4))

    for col in sheet.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[col_letter].width = max_length + 2

    workbook.save(output_path)
    print(f"\nResults saved to: {output_path}")


# ============================================================================
# PART 2 — SPRING CONSTANT CALCULATION FUNCTIONS
# ============================================================================


def compute_spring_constants(spreadsheet_path: str, scale: float, weight_mass: float, weight_num: np.ndarray):
    """
    Read Y displacement data from the xlsx produced by Part 1, convert to meters,
    and compute the spring constant via linear regression of force vs. displacement.

    Returns:
        spring_constant: float (N/m)
        r_squared: float
        displacement_m: np.ndarray — displacements in meters
        force: np.ndarray — applied forces in Newtons
        regression_result: linregress result object
    """
    force = weight_num * weight_mass * GRAVITY

    wb = openpyxl.load_workbook(spreadsheet_path, data_only=True)
    ws = wb.active

    # Read Y displacement values from column C (column 3, skip header row)
    y_disp_pixels = []
    for row in range(2, 2 + len(weight_num)):
        cell_value = ws.cell(row=row, column=3).value
        if cell_value is not None:
            y_disp_pixels.append(float(cell_value))

    y_disp_pixels = np.array(y_disp_pixels)

    # Convert pixel displacement to meters
    displacement_m = y_disp_pixels / scale

    result = linregress(force, displacement_m)
    spring_constant = 1.0 / result.slope if result.slope != 0 else float("inf")
    r_squared = result.rvalue**2

    return spring_constant, r_squared, displacement_m, force, result


def plot_spring_constant_results(force, displacement_m, regression_result, spring_constant, r_squared):
    """
    Plot force vs. displacement with regression line.
    """
    plt.figure()
    plt.plot(force, displacement_m, "D", label="Measured")
    plt.plot(force, regression_result.slope * force + regression_result.intercept, label="Linear fit")
    plt.ylabel("Post head displacement (m)")
    plt.xlabel("Applied force (N)")
    plt.title(f"Spring constant: {spring_constant:.4f} N/m  (R\u00b2 = {r_squared:.6f})")
    plt.legend()
    plt.grid(True)
    plt.tight_layout()
    plt.show()


# ============================================================================
# MAIN
# ============================================================================

if __name__ == "__main__":
    # --- Part 1: Image tracking ---

    print("loading image dir")
    if IMAGE_DIR is None:
        import platform
        if platform.system() == "Darwin":
            import subprocess
            result = subprocess.run(
                ["osascript", "-e", 'POSIX path of (choose folder with prompt "Select folder containing images to analyze")'],
                capture_output=True, text=True,
            )
            selected_dir = result.stdout.strip()
        else:
            from tkinter import Tk, filedialog
            root = Tk()
            root.withdraw()
            selected_dir = filedialog.askdirectory(title="Select folder containing images to analyze")
            root.destroy()
        if not selected_dir:
            print("ERROR: No folder selected. Exiting.")
            sys.exit(1)
        IMAGE_DIR = selected_dir

    if not os.path.isdir(IMAGE_DIR):
        print(f"ERROR: '{IMAGE_DIR}' is not a valid directory.")
        sys.exit(1)

    output_xlsx = OUTPUT_DIR
    if output_xlsx is None:
        output_xlsx = os.path.join(Path(IMAGE_DIR).parent, f"{Path(IMAGE_DIR).name}_deflection_results.xlsx")

    print(f"Loading images from: {IMAGE_DIR}")
    images = load_images(IMAGE_DIR)
    print(f"Found {len(images)} images.\n")

    # --- Scale calibration ---
    if SCALE is None:
        scale_px_per_m = calibrate_scale(images[0][1], FEATURE_DIAMETER_MM)
    else:
        scale_px_per_m = SCALE * 1000.0  # convert px/mm to px/m
        print(f"  Using manual scale: {SCALE} px/mm ({scale_px_per_m} px/m)")

    # --- Template selection and tracking ---
    template_gray, roi = select_template(images[0][1])

    max_rotation = MAX_ROTATION_DEFAULT
    if max_rotation is not None:
        print(f"Rotation search enabled: +/- {max_rotation} degrees in {ROTATION_INCREMENT} degree steps")

    print("\nTracking template across all images...")
    results = track_across_images(images, template_gray, max_rotation=max_rotation)

    print("\nSaving annotated images...")
    annotated_dir = os.path.join(IMAGE_DIR, "tracked")
    save_annotated_images(results, annotated_dir)

    write_results_to_xlsx(results, output_xlsx)

    # --- Part 2: Spring constant calculation ---

    print("\n" + "=" * 60)
    print("SPRING CONSTANT CALCULATION")
    print("=" * 60)
    print(f"  Scale: {scale_px_per_m:.1f} pixels/m")
    print(f"  Weight mass: {WEIGHT_MASS} kg")
    print(f"  Number of weight levels: {len(WEIGHT_NUM)}")

    spring_constant, r_squared, displacement_m, force, regression = compute_spring_constants(
        output_xlsx, scale_px_per_m, WEIGHT_MASS, WEIGHT_NUM
    )

    print(f"\n  Spring constant: {spring_constant:.4f} N/m")
    print(f"  R\u00b2: {r_squared:.6f}")

    plot_spring_constant_results(force, displacement_m, regression, spring_constant, r_squared)

    print("\nDone.")
